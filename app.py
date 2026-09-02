"""
Employee Attendance Tracker (with Admin Panel)
------------------------------------------------
A Flask web app that records employee attendance (login, breaks, lunch,
logout, meeting details) and automatically saves every entry into an
Excel file. A NEW SHEET is created automatically for each date.

Admin panel (login required):
    - View any date's records
    - Edit or delete a record
    - Download the full Excel workbook

Run it with:
    pip install flask openpyxl --break-system-packages
    python app.py

Then open http://127.0.0.1:5000 in your browser.
Admin login: http://127.0.0.1:5000/admin/login

Excel file created at: attendance_data/attendance.xlsx

SECURITY NOTE: Admin credentials are set below in ADMIN_USERNAME /
ADMIN_PASSWORD. This demo checks them in plain text, which is fine for
a small internal/local tool, but if you ever deploy this somewhere
public, swap in hashed passwords and HTTPS.
"""

import os
from datetime import datetime
from functools import wraps

from flask import (
    Flask, render_template_string, request, redirect, url_for,
    flash, session, send_file, abort
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
# Secret key & admin credentials now come from environment variables so real
# values are never committed to source control. Falls back to a dev default
# only when the env var isn't set (e.g. running locally).
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DATA_FOLDER = "attendance_data"
EXCEL_FILE = os.path.join(DATA_FOLDER, "attendance.xlsx")

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "Mobius365")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Mobius@123")

COLUMNS = [
    "Emp_Name",
    "Emp_ID",
    "Login Time",
    "Break_1",
    "Lunch",
    "Break_2",
    "Log Out Time",
    "Meeting Hr/Details",
]

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
CELL_FONT = Font(name="Arial")


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def ensure_workbook():
    """Create the workbook + folder if they don't already exist."""
    os.makedirs(DATA_FOLDER, exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.active.title = "Info"
        wb.active["A1"] = "Employee Attendance Tracker — data is stored in date-named sheets."
        wb.save(EXCEL_FILE)


def style_header(ws):
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 12, 14, 12, 12, 12, 14, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def get_or_create_sheet(wb, sheet_name):
    """Return the sheet for a given date, creating it (styled) if needed."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)
    style_header(ws)
    return ws


def save_entry(data: dict):
    """Append one attendance row to today's sheet in the Excel workbook."""
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    today_str = datetime.now().strftime("%Y-%m-%d")
    ws = get_or_create_sheet(wb, today_str)

    row = [data.get(col, "") for col in COLUMNS]
    ws.append(row)

    new_row_idx = ws.max_row
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=new_row_idx, column=col_idx).font = CELL_FONT

    wb.save(EXCEL_FILE)


def get_records_for_date(date_str):
    """Read rows (with their Excel row number) for a given date's sheet."""
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE, data_only=True)
    if date_str not in wb.sheetnames:
        return []
    ws = wb[date_str]
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and any(row):
            record = dict(zip(COLUMNS, row))
            record["_row"] = row_idx
            records.append(record)
    return records


def get_today_records():
    return get_records_for_date(datetime.now().strftime("%Y-%m-%d"))


def list_available_dates():
    """List every date-named sheet, most recent first."""
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE, data_only=True)
    dates = [name for name in wb.sheetnames if name != "Info"]
    dates.sort(reverse=True)
    return dates


def update_record(date_str, row_idx, data: dict):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    if date_str not in wb.sheetnames:
        return False
    ws = wb[date_str]
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=data.get(col, ""))
        cell.font = CELL_FONT
    wb.save(EXCEL_FILE)
    return True


def delete_record(date_str, row_idx):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    if date_str not in wb.sheetnames:
        return False
    ws = wb[date_str]
    ws.delete_rows(row_idx, 1)
    wb.save(EXCEL_FILE)
    return True


# ---------------------------------------------------------------------------
# Admin auth helper
# ---------------------------------------------------------------------------
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)
    return wrapper


# ---------------------------------------------------------------------------
# HTML templates
# ---------------------------------------------------------------------------
BASE_STYLE = """
<style>
    body { font-family: Arial, sans-serif; max-width: 900px; margin: 30px auto; background: #f5f6fa; color: #222; }
    h1 { color: #2c3e50; }
    .card { background: #fff; padding: 24px; border-radius: 10px; box-shadow: 0 2px 6px rgba(0,0,0,0.08); margin-bottom: 24px; }
    label { display: block; margin-top: 12px; font-weight: bold; font-size: 14px; }
    input[type=text], input[type=time], input[type=password], select {
        width: 100%; padding: 8px; margin-top: 4px; border: 1px solid #ccc; border-radius: 6px; box-sizing: border-box;
    }
    button, .btn { margin-top: 18px; padding: 10px 22px; background: #305496; color: white; border: none;
        border-radius: 6px; font-size: 15px; cursor: pointer; text-decoration: none; display: inline-block; }
    button:hover, .btn:hover { background: #24406f; }
    .btn-danger { background: #b23b3b; }
    .btn-danger:hover { background: #8f2e2e; }
    .btn-small { padding: 5px 12px; font-size: 13px; margin: 0 4px 0 0; }
    table { width: 100%; border-collapse: collapse; margin-top: 10px; }
    th, td { border: 1px solid #ddd; padding: 8px; font-size: 13px; text-align: left; }
    th { background: #305496; color: white; }
    .flash { padding: 10px 14px; background: #d4edda; color: #155724; border-radius: 6px; margin-bottom: 16px; }
    .flash-error { background: #f8d7da; color: #721c24; }
    .note { font-size: 13px; color: #555; }
    .topbar { display: flex; justify-content: space-between; align-items: center; }
    .topbar a { color: #305496; text-decoration: none; font-size: 14px; }
</style>
"""

PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Employee Attendance Tracker</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>🕒 Employee Attendance Tracker</h1>
        <a href="{{ url_for('admin_login') }}">🔐 Admin Login</a>
    </div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}<div class="flash">{{ m }}</div>{% endfor %}
      {% endif %}
    {% endwith %}

    <div class="card">
        <form method="POST" action="{{ url_for('submit') }}">
            <label>Emp_Name</label>
            <input type="text" name="Emp_Name" required>
            <label>Emp_ID</label>
            <input type="text" name="Emp_ID" required>
            <label>Login Time</label>
            <input type="time" name="Login Time">
            <label>Break_1</label>
            <input type="time" name="Break_1">
            <label>Lunch</label>
            <input type="time" name="Lunch">
            <label>Break_2</label>
            <input type="time" name="Break_2">
            <label>Log Out Time</label>
            <input type="time" name="Log Out Time">
            <label>Meeting Hr/Details</label>
            <input type="text" name="Meeting Hr/Details" placeholder="e.g. 2:00 PM - Team sync">
            <button type="submit">Save Entry</button>
        </form>
        <p class="note">Every submission is saved automatically to <code>attendance_data/attendance.xlsx</code>, on a sheet named for today's date ({{ today }}).</p>
    </div>

    <div class="card">
        <h2>Today's Records — {{ today }}</h2>
        {% if records %}
        <table>
            <tr>{% for col in columns %}<th>{{ col }}</th>{% endfor %}</tr>
            {% for r in records %}
            <tr>{% for col in columns %}<td>{{ r[col] }}</td>{% endfor %}</tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No entries yet today.</p>
        {% endif %}
    </div>
</body>
</html>
"""

LOGIN_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Admin Login</title>""" + BASE_STYLE + """</head>
<body>
    <h1>🔐 Admin Login</h1>
    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}<div class="flash flash-error">{{ m }}</div>{% endfor %}
      {% endif %}
    {% endwith %}
    <div class="card">
        <form method="POST">
            <label>Username</label>
            <input type="text" name="username" required autofocus>
            <label>Password</label>
            <input type="password" name="password" required>
            <button type="submit">Login</button>
        </form>
    </div>
    <p><a href="{{ url_for('home') }}">← Back to attendance form</a></p>
</body>
</html>
"""

ADMIN_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Admin Panel</title>""" + BASE_STYLE + """</head>
<body>
    <div class="topbar">
        <h1>🛠 Admin Panel</h1>
        <a href="{{ url_for('admin_logout') }}">Log out</a>
    </div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}<div class="flash">{{ m }}</div>{% endfor %}
      {% endif %}
    {% endwith %}

    <div class="card">
        <form method="GET" action="{{ url_for('admin_panel') }}">
            <label>Select date</label>
            <select name="date" onchange="this.form.submit()">
                {% for d in dates %}
                <option value="{{ d }}" {% if d == selected_date %}selected{% endif %}>{{ d }}</option>
                {% endfor %}
            </select>
        </form>
        <a class="btn" href="{{ url_for('download_excel') }}">⬇ Download Full Excel File</a>
    </div>

    <div class="card">
        <h2>Records — {{ selected_date }}</h2>
        {% if records %}
        <table>
            <tr>
                {% for col in columns %}<th>{{ col }}</th>{% endfor %}
                <th>Actions</th>
            </tr>
            {% for r in records %}
            <tr>
                {% for col in columns %}<td>{{ r[col] }}</td>{% endfor %}
                <td>
                    <a class="btn btn-small" href="{{ url_for('admin_edit', date=selected_date, row=r['_row']) }}">Edit</a>
                    <form style="display:inline" method="POST" action="{{ url_for('admin_delete', date=selected_date, row=r['_row']) }}"
                          onsubmit="return confirm('Delete this record?');">
                        <button class="btn btn-small btn-danger" type="submit">Delete</button>
                    </form>
                </td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
        <p>No records for this date.</p>
        {% endif %}
    </div>
</body>
</html>
"""

EDIT_PAGE = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>Edit Record</title>""" + BASE_STYLE + """</head>
<body>
    <h1>✏️ Edit Record — {{ date }}</h1>
    <div class="card">
        <form method="POST">
            <label>Emp_Name</label>
            <input type="text" name="Emp_Name" value="{{ record['Emp_Name'] or '' }}" required>
            <label>Emp_ID</label>
            <input type="text" name="Emp_ID" value="{{ record['Emp_ID'] or '' }}" required>
            <label>Login Time</label>
            <input type="text" name="Login Time" value="{{ record['Login Time'] or '' }}">
            <label>Break_1</label>
            <input type="text" name="Break_1" value="{{ record['Break_1'] or '' }}">
            <label>Lunch</label>
            <input type="text" name="Lunch" value="{{ record['Lunch'] or '' }}">
            <label>Break_2</label>
            <input type="text" name="Break_2" value="{{ record['Break_2'] or '' }}">
            <label>Log Out Time</label>
            <input type="text" name="Log Out Time" value="{{ record['Log Out Time'] or '' }}">
            <label>Meeting Hr/Details</label>
            <input type="text" name="Meeting Hr/Details" value="{{ record['Meeting Hr/Details'] or '' }}">
            <button type="submit">Save Changes</button>
        </form>
    </div>
    <p><a href="{{ url_for('admin_panel', date=date) }}">← Back to admin panel</a></p>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Public routes
# ---------------------------------------------------------------------------
@app.route("/")
def home():
    records = get_today_records()
    today = datetime.now().strftime("%Y-%m-%d")
    return render_template_string(PAGE, records=records, columns=COLUMNS, today=today)


@app.route("/submit", methods=["POST"])
def submit():
    data = {col: request.form.get(col, "").strip() for col in COLUMNS}
    save_entry(data)
    flash(f"Saved entry for {data['Emp_Name']} ({data['Emp_ID']}).")
    return redirect(url_for("home"))


# ---------------------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Logged in successfully.")
            return redirect(url_for("admin_panel"))
        flash("Invalid username or password.")
    return render_template_string(LOGIN_PAGE)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("home"))


@app.route("/admin")
@admin_required
def admin_panel():
    ensure_workbook()
    dates = list_available_dates()
    today = datetime.now().strftime("%Y-%m-%d")
    selected_date = request.args.get("date") or (dates[0] if dates else today)
    records = get_records_for_date(selected_date)
    return render_template_string(
        ADMIN_PAGE, dates=dates, selected_date=selected_date,
        records=records, columns=COLUMNS
    )


@app.route("/admin/edit/<date>/<int:row>", methods=["GET", "POST"])
@admin_required
def admin_edit(date, row):
    if request.method == "POST":
        data = {col: request.form.get(col, "").strip() for col in COLUMNS}
        if update_record(date, row, data):
            flash("Record updated.")
        else:
            flash("Could not update record — sheet not found.")
        return redirect(url_for("admin_panel", date=date))

    records = get_records_for_date(date)
    record = next((r for r in records if r["_row"] == row), None)
    if record is None:
        abort(404)
    return render_template_string(EDIT_PAGE, date=date, record=record)


@app.route("/admin/delete/<date>/<int:row>", methods=["POST"])
@admin_required
def admin_delete(date, row):
    if delete_record(date, row):
        flash("Record deleted.")
    else:
        flash("Could not delete record — sheet not found.")
    return redirect(url_for("admin_panel", date=date))


@app.route("/admin/download")
@admin_required
def download_excel():
    ensure_workbook()
    return send_file(EXCEL_FILE, as_attachment=True, download_name="attendance.xlsx")


ensure_workbook()

if __name__ == "__main__":
    # host="0.0.0.0" makes the app reachable from other devices on the same
    # WiFi network, not just this computer. Debug is off for that reason too
    # (the debugger would otherwise be exposed to the network).
    # PORT is read from the environment because Render (and most hosts)
    # assign the port dynamically rather than always using 5000.
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
